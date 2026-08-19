from collections import Counter, defaultdict
from pathlib import Path
from secrets import token_urlsafe

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.utils.text import slugify
from openpyxl import load_workbook

from scoring.models import EventCategory, Game, Player, ScorekeeperAccount, Team
from scoring.team_utils import player_key, team_key


TEAM_EVENTS = [
    "Paper Floor Relay",
    "Spin Transfer",
    "Flip It - Ping Pong Ball into cup",
    "Drop the bottle (finish line)",
    "Balloon Bowling",
    "Ping Pong Ball Relay",
    "Crossfire",
    "Catapult",
    "Pick up the tab (straw and cans)",
    "Balloon Pyramid Game",
]

INDIVIDUAL_EVENTS = [
    "Rubber Band Archery",
    "Cup Unstack",
    "Binder Clip Flick",
    "Water Slide",
    "Sticky Note Hi Tag",
    "Fan the cup challenge",
    "Pencil Javelin",
    "Paper Plane Long Jump",
]

DOUBLES_EVENTS = [
    "Balloon Race",
    "Swap the colour",
    "Not my hand",
    "Chopstick Ball Transfer",
    "Chair Race",
    "Office Tennis",
    "Cup Catch",
    "Marshmallow Catch",
]


class Command(BaseCommand):
    help = "Seed Office Olympics teams, players, games, and optional local accounts."

    def add_arguments(self, parser):
        parser.add_argument("--excel", type=Path, help="Path to Office Olympics workbook.")
        parser.add_argument("--create-admin", action="store_true", help="Create admin/admin123 if no admin exists.")
        parser.add_argument(
            "--create-scorekeepers",
            action="store_true",
            help="Create one random scorekeeper account per game and print credentials.",
        )

    def handle(self, *args, **options):
        self.seed_games()
        if options["excel"]:
            self.import_workbook(options["excel"])
        if options["create_admin"]:
            self.create_admin()
        if options["create_scorekeepers"]:
            self.create_scorekeepers()
        self.stdout.write(self.style.SUCCESS("Office Olympics seed complete."))

    def seed_games(self):
        groups = [
            (EventCategory.TEAM, TEAM_EVENTS),
            (EventCategory.INDIVIDUAL, INDIVIDUAL_EVENTS),
            (EventCategory.DOUBLES, DOUBLES_EVENTS),
        ]
        for category, names in groups:
            for index, name in enumerate(names, start=1):
                Game.objects.update_or_create(
                    name=name,
                    defaults={"category": category, "sort_order": index, "active": True},
                )

    def import_workbook(self, path):
        if not path.exists():
            raise CommandError(f"Workbook not found: {path}")
        workbook = load_workbook(path, data_only=True)

        shirt_data = defaultdict(lambda: {"size": Counter(), "colour": Counter()})
        if "Name Size Colour" in workbook.sheetnames:
            for name, size, colour in workbook["Name Size Colour"].iter_rows(min_row=2, values_only=True):
                if not name:
                    continue
                key = str(name).strip()
                if size:
                    shirt_data[key]["size"][str(size).strip()] += 1
                if colour:
                    shirt_data[key]["colour"][str(colour).strip()] += 1

        if "Names by Team" not in workbook.sheetnames:
            raise CommandError("Workbook must contain a 'Names by Team' sheet.")

        seen_pairs = set()
        for team_name, player_name in workbook["Names by Team"].iter_rows(min_row=2, values_only=True):
            if not team_name or not player_name:
                continue
            team_name = str(team_name).strip()
            player_name = str(player_name).strip()
            pair = (team_name.lower(), player_name.lower())
            if pair in seen_pairs:
                continue
            seen_pairs.add(pair)

            data = shirt_data[player_name]
            shirt_size = data["size"].most_common(1)[0][0] if data["size"] else ""
            shirt_colour = data["colour"].most_common(1)[0][0] if data["colour"] else ""
            team = next((existing for existing in Team.objects.all() if team_key(existing.name) == team_key(team_name)), None)
            if team is None:
                team = Team.objects.create(name=team_name, colour=shirt_colour)
            if shirt_colour and not team.colour:
                team.colour = shirt_colour
                team.save(update_fields=["colour"])
            player = next(
                (
                    existing
                    for existing in Player.objects.filter(team=team)
                    if player_key(existing.name) == player_key(player_name)
                ),
                None,
            )
            if player is None:
                Player.objects.create(name=player_name, team=team, shirt_size=shirt_size, shirt_colour=shirt_colour)
            else:
                updates = []
                if shirt_size and not player.shirt_size:
                    player.shirt_size = shirt_size
                    updates.append("shirt_size")
                if shirt_colour and not player.shirt_colour:
                    player.shirt_colour = shirt_colour
                    updates.append("shirt_colour")
                if updates:
                    player.save(update_fields=updates)

    def create_admin(self):
        User = get_user_model()
        if User.objects.filter(is_superuser=True).exists():
            return
        User.objects.create_superuser(username="admin", email="", password="admin123")
        self.stdout.write(self.style.WARNING("Created Django admin: admin / admin123"))

    def create_scorekeepers(self):
        for game in Game.objects.order_by("category", "sort_order", "name"):
            username = slugify(game.name)[:48] or f"game-{game.pk}"
            password = token_urlsafe(8)
            account, created = ScorekeeperAccount.objects.get_or_create(username=username)
            if created:
                account.set_password(password)
                account.save()
                account.games.add(game)
                self.stdout.write(f"{game.name}: {username} / {password}")
